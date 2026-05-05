# -*- coding: utf-8 -*-
"""
CDAP清单 Excel 转 md 文件 - v2
改进分区字段检测逻辑，从清单目录获取视图名和负责人
"""
import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 配置
EXCEL_FILE = 'CDAP自助分析-清单目录_修正版2.xlsx'
OUTPUT_DIR = '.claude/skills/write-query/references/tables'

# 读取 Excel
wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

# 读取清单目录，构建表索引
index_sheet = wb[wb.sheetnames[0]]
index_rows = list(index_sheet.iter_rows(values_only=True))

# 构建索引: 表名 -> {seq, hive_table, view_name, business_owner, tech_owner}
# 同时用 Hive表名 也建立索引（某些sheet名可能是Hive表名格式）
table_info_by_name = {}
table_info_by_hive = {}

for row in index_rows[1:]:
    if row[0] is None:
        continue
    seq = row[0]
    name = row[1] if row[1] else ''
    hive_table = row[2] if row[2] else ''
    view_name = row[3] if row[3] else ''
    business_owner = row[4] if row[4] else ''
    tech_owner = row[5] if row[5] else ''

    info = {
        'seq': seq,
        'name': name,
        'hive_table': hive_table,
        'view_name': view_name,
        'business_owner': business_owner,
        'tech_owner': tech_owner
    }

    if name:
        table_info_by_name[name] = info
    if hive_table:
        table_info_by_hive[hive_table] = info

print(f"Loaded {len(table_info_by_name)} tables from index")

def normalize_name(name):
    """标准化表名，便于匹配"""
    # 去掉括号内容
    name = name.split('(')[0].strip()
    # 常见缩写映射
    replacements = {
        '酒宽': '酒店宽带',
        '主宽': '宽带',
        '主宽拆机挽留': '宽带拆机挽留',
    }
    for old, new in replacements.items():
        if old in name:
            name = name.replace(old, new)
    return name

def find_table_info(sheet_name):
    """根据sheet名查找表信息"""
    # 去掉括号部分并标准化
    name_from_sheet = normalize_name(sheet_name)

    # 1. 精确匹配 sheet名（标准化后）
    if name_from_sheet in table_info_by_name:
        return table_info_by_name[name_from_sheet]

    # 2. 模糊匹配（sheet名包含在表名中，或表名包含sheet名）
    for name, info in table_info_by_name.items():
        normalized_name = normalize_name(name)
        if name_from_sheet in normalized_name or normalized_name in name_from_sheet:
            return info
        # 去掉"（已停用）"后匹配
        normalized_name_no_suffix = normalized_name.replace('（已停用）', '').replace('(已停用)', '')
        if name_from_sheet in normalized_name_no_suffix or normalized_name_no_suffix in name_from_sheet:
            return info

    # 3. 用 Hive表名 匹配（某些sheet可能是 hive表名格式）
    if '(' in sheet_name:
        hive_from_sheet = sheet_name.split('(')[1].rstrip(')')
        if hive_from_sheet in table_info_by_hive:
            return table_info_by_hive[hive_from_sheet]

    return None

def find_partition_field(rows):
    """查找分区字段 - 检查所有单元格，优先 par_month_id"""
    partition_candidates = []

    for row in rows[2:]:  # 跳过表名行和表头行
        for cell in row:
            if cell and 'par_month' in str(cell).lower():
                partition_candidates.append(str(cell))

    # 优先返回 par_month_id
    for p in partition_candidates:
        if p == 'par_month_id':
            return p

    return partition_candidates[0] if partition_candidates else ''

# 遍历所有 sheet（跳过清单目录）
converted_count = 0
error_count = 0

for sheet_name in wb.sheetnames[1:]:  # skip first sheet (清单目录)
    try:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 3:
            print(f"  Skip {sheet_name}: only {len(rows)} rows")
            continue

        # 从 sheet 名提取表名
        if '(' in sheet_name:
            table_name = sheet_name.split('(')[0].strip()
        else:
            table_name = sheet_name.strip()

        # 查找表信息
        info = find_table_info(sheet_name)
        if info:
            seq = info['seq']
            hive_table = info['hive_table']
            view_name = info['view_name']
            business_owner = info['business_owner']
            tech_owner = info['tech_owner']
        else:
            seq = 0
            hive_table = rows[0][0] if rows[0][0] else ''
            view_name = '待补充'
            business_owner = ''
            tech_owner = ''

        # 提取分区字段
        partition_field = find_partition_field(rows)

        # 生成 md 内容
        md_lines = []
        md_lines.append(f'# {table_name}')
        md_lines.append('')
        md_lines.append(f'- **Hive 表名**: `{hive_table}`')
        md_lines.append(f'- **视图名**: {view_name}')
        md_lines.append(f'- **业务负责人**: {business_owner}')
        md_lines.append(f'- **业支负责人**: {tech_owner}')
        if partition_field:
            md_lines.append(f'- **分区字段**: {partition_field}')
        md_lines.append('')
        md_lines.append('---')
        md_lines.append('')
        md_lines.append('## 字段说明')
        md_lines.append('')
        md_lines.append('| 字段 | 字段类型 | 字段含义 |')
        md_lines.append('|------|---------|---------|')

        # 检测数据格式：
        # 格式A: col0=字段名, col1=类型, col2=含义（大多数表）
        # 格式B: col0=None, col1=字段名, col2=类型, col3=含义（全业务资料表等）
        # 通过检查前几行数据来判断
        has_col0_data = False
        has_col1_data = False
        for row in rows[2:6]:  # 检查前几行
            if row[0] is not None and row[1] is not None:
                has_col0_data = True
            if row[1] is not None and row[2] is not None:
                has_col1_data = True

        use_format_b = has_col1_data and not has_col0_data

        # 跳过表名行和表头行，从第3行开始
        for row in rows[2:]:
            if use_format_b:
                # 格式B: col0=None, col1=字段名, col2=类型, col3=含义
                if row[1] is not None:
                    field_name = row[1] if row[1] else ''
                    field_type = row[2] if row[2] else ''
                    field_desc = row[3] if row[3] else ''
                    md_lines.append(f'| {field_name} | {field_type} | {field_desc} |')
            else:
                # 格式A: col0=字段名, col1=类型, col2=含义
                if row[0] is not None:
                    field_name = row[0] if row[0] else ''
                    field_type = row[1] if row[1] else ''
                    field_desc = row[2] if row[2] else ''
                    md_lines.append(f'| {field_name} | {field_type} | {field_desc} |')

        md_content = '\n'.join(md_lines)

        # 写入文件
        filename = f'{seq:03d}_{table_name}.md'
        filepath = os.path.join(OUTPUT_DIR, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(md_content)

        converted_count += 1
        status = ''
        if not info:
            status = ' [NOT FOUND IN INDEX]'
        print(f"  [OK] {filename}{status}")

    except Exception as e:
        error_count += 1
        print(f"  [ERROR] {sheet_name}: {e}")

print(f"\nDone: {converted_count} converted, {error_count} errors")
