# -*- coding: utf-8 -*-
"""
CDAP清单 Excel 转 md 文件
将每个 sheet 转换为独立的表结构 md 文档
"""
import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

# 配置
EXCEL_FILE = 'CDAP自助分析-清单目录_修正版2.xlsx'
OUTPUT_DIR = '.claude/skills/write-query/references/tables'

# 读取 Excel
wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True, data_only=True)

# 读取清单目录，构建表索引
index_sheet = wb[wb.sheetnames[0]]
index_rows = list(index_sheet.iter_rows(values_only=True))

# 构建索引: sheet名 -> {hive_table, view_name, business_owner, tech_owner}
table_info = {}
for row in index_rows[1:]:
    if row[0] is None:
        continue
    seq = row[0]
    name = row[1] if row[1] else ''
    hive_table = row[2] if row[2] else ''
    view_name = row[3] if row[3] else ''
    business_owner = row[4] if row[4] else ''
    tech_owner = row[5] if row[5] else ''

    # 用 name 作为 key 匹配 sheet
    if name:
        # sheet 名可能是 "降档清单(ads_yz_jd_list)" 格式
        table_info[name] = {
            'seq': seq,
            'name': name,
            'hive_table': hive_table,
            'view_name': view_name,
            'business_owner': business_owner,
            'tech_owner': tech_owner
        }

print(f"Loaded {len(table_info)} tables from index")

# 遍历所有 sheet（跳过清单目录）
converted_count = 0
error_count = 0

for sheet_name in wb.sheetnames[1:]:  # skip first sheet (清单目录)
    try:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 3:
            print(f"  Skip {sheet_name}: only {len(rows)} rows")
            continue

        # 从 sheet 名提取表名（去掉括号部分）
        # sheet_name 格式: "降档清单(ads_yz_jd_list)" 或 "全业务资料表"
        if '(' in sheet_name:
            table_name = sheet_name.split('(')[0].strip()
        else:
            table_name = sheet_name.strip()

        # 查找表信息
        info = table_info.get(table_name, {})
        seq = info.get('seq', 0)
        hive_table = info.get('hive_table', rows[0][0] if rows[0][0] else '')
        view_name = info.get('view_name', '待补充')
        business_owner = info.get('business_owner', '')
        tech_owner = info.get('tech_owner', '')

        # 提取分区字段（查找 par_month_id 或类似字段）
        partition_field = ''
        for row in rows[2:]:
            if row[0] and 'par_month' in str(row[0]).lower():
                partition_field = row[0]
                break

        # 生成 md 内容
        md_lines = []
        md_lines.append(f'# {table_name}')
        md_lines.append('')
        md_lines.append(f'- **Hive 表名**: `{hive_table}`')
        md_lines.append(f'- **视图名**: {view_name}')
        md_lines.append(f'- **业务负责人**: {business_owner}')
        md_lines.append(f'- **业支负责人**: {tech_owner}')
        if partition_field:
            md_lines.append(f'- **分区字段**: {partition_field}')
        md_lines.append('')
        md_lines.append('---')
        md_lines.append('')
        md_lines.append('## 字段说明')
        md_lines.append('')
        md_lines.append('| 字段 | 字段类型 | 字段含义 |')
        md_lines.append('|------|---------|---------|')

        # 跳过表名行和表头行，从第3行开始
        for row in rows[2:]:
            if row[0] is not None:
                field_name = row[0] if row[0] else ''
                field_type = row[1] if row[1] else ''
                field_desc = row[2] if row[2] else ''
                md_lines.append(f'| {field_name} | {field_type} | {field_desc} |')

        md_content = '\n'.join(md_lines)

        # 写入文件
        filename = f'{seq:03d}_{table_name}.md'
        filepath = os.path.join(OUTPUT_DIR, filename)
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(md_content)

        converted_count += 1
        print(f"  [OK] {filename}")

    except Exception as e:
        error_count += 1
        print(f"  [ERROR] {sheet_name}: {e}")

print(f"\nDone: {converted_count} converted, {error_count} errors")
