#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel_to_table_md.py

将 CDAP 清单 Excel 转换为 write-query 技能规范的表结构 md 文档。

处理流程：
1. 使用 openpyxl 读取 Excel 文件
2. 解析并提取：表名、Hive表名、视图名、字段分类、字段详情、口径案例
3. 按模板格式重组，输出到技能 references 目录
"""

import argparse
import os
import re
import sys
import logging

import openpyxl

# 路径配置
SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
REFERENCES_DIR = os.path.join(SKILL_DIR, "references")


# 解决 Windows GBK 编码问题
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
    stream=sys.stdout
)
logger = logging.getLogger("excel_to_table_md")


def read_excel_to_md(excel_path: str) -> str:
    """
    使用 openpyxl 读取 Excel 文件，生成中间 md 格式。
    中间 md 格式兼容原有 parse_intermediate_md() 的解析逻辑。
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    lines = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"## {sheet_name}\n")

        for row in ws.iter_rows(values_only=True):
            # 跳过全空行
            if not any(cell is not None for cell in row):
                continue

            # 将行转换为 md 表格格式（取前7列）
            cells = [str(cell) if cell is not None else "" for cell in row[:7]]
            # 清理单元格内容中的管道符
            cells = [c.replace("|", "\\|") for c in cells]
            lines.append("| " + " | ".join(cells) + " |")

    return "\n".join(lines)


def parse_intermediate_md(content: str) -> dict:
    """
    解析中间 md，提取关键信息。

    返回结构：
    {
        "table_name": str,          # 中文表名
        "hive_table": str,           # Hive 表名
        "view_name": str,            # 视图名
        "sections": [                # 字段分类 sections
            {
                "category": str,     # 分类名称（如"新入网"，空字符串表示"基础字段"）
                "fields": [          # 该分类下的字段列表
                    {
                        "field": str,
                        "meaning": str,
                        "cycle": str,
                        "dict_value": str,
                        "note": str
                    }
                ]
            }
        ],
        "case_metrics": [            # 口径案例
            {"name": str, "sql": str}
        ]
    }
    支持两种 Excel 格式：
    - 069 格式：有字段分类列，6列（分类/字段/含义/周期/字典值/说明）
    - 002 格式：无分类列，只有3列（字段名/字段含义/注释）
    """
    lines = content.split("\n")

    result = {
        "table_name": "",
        "hive_table": "",
        "view_name": "",
        "sections": [],
        "case_metrics": []
    }

    # 提取表名（中文名）：从第一行的 ## {sheet_name} 中提取
    for line in lines:
        line = line.strip()
        m = re.match(r"^##\s+(.+)$", line)
        if m:
            result["table_name"] = m.group(1).strip()
            break

    # 提取 Hive表名、视图名、英文表名（从表格行中提取）
    for line in lines:
        line = line.strip()
        if not line.startswith("|"):
            continue
        cells = [c.strip() for c in line.split("|")[1:-1]]
        if len(cells) < 2:
            continue

        # 精确匹配：避免 "Hive" in "表名：" 导致误匹配
        if re.fullmatch(r"Hive表名[：:]?", cells[0]):
            result["hive_table"] = cells[1] if len(cells) > 1 else ""
        elif re.fullmatch(r"视图名[：:]?", cells[0]):
            result["view_name"] = cells[1] if len(cells) > 1 else ""
        elif re.fullmatch(r"表名[：:]?", cells[0]):
            # "表名："行通常是英文 Hive 表名（如 ads_yz_tb_comm_cm_all_final）
            if not result["hive_table"] and len(cells) > 1 and cells[1]:
                result["hive_table"] = cells[1]

    # 判断 Excel 格式类型
    has_category_column = False
    for line in lines:
        if re.search(r"字段分类", line):
            has_category_column = True
            break

    if has_category_column:
        # === 069 格式：有字段分类列 ===
        _parse_sections_with_category(lines, result)
    else:
        # === 002 格式：无分类列，只有字段名/字段含义/注释 ===
        _parse_sections_without_category(lines, result)

    return result


def _parse_sections_with_category(lines: list, result: dict):
    """解析有字段分类的格式（069 类型）"""
    # 找到字段表头行（包含"字段分类"和"字段含义"的行）
    header_line_idx = -1
    for i, line in enumerate(lines):
        if re.search(r"字段分类", line) and re.search(r"字段含义", line):
            header_line_idx = i
            break

    if header_line_idx == -1:
        logger.warning("未找到字段表头，跳过字段解析")
        return

    # 解析字段分类
    current_category = ""
    current_fields = []

    # 从表头下一行开始解析
    for i in range(header_line_idx + 1, len(lines)):
        line = lines[i].strip()

        # 跳过空行和分隔线
        if not line or line.startswith("| ---"):
            continue

        # 跳过非数据行
        if not line.startswith("|"):
            continue

        # 遇到案例指标标题，停止字段解析
        if re.search(r"案例指标|语句", line) and "|" not in line.split("|")[1]:
            if re.search(r"案例指标|口径|指标名称", line):
                break

        cells = [c.strip() for c in line.split("|")[1:-1]]

        if len(cells) < 2:
            continue

        # 字段分类列有内容，表示遇到了新的分类
        category_cell = cells[0] if len(cells) > 0 else ""
        if category_cell and category_cell not in ("", "字段分类"):
            if current_fields or current_category:
                result["sections"].append({
                    "category": current_category or "基础字段",
                    "fields": current_fields
                })
            current_category = category_cell
            current_fields = []

            if len(cells) < 2 or not cells[1]:
                continue

        field_name = cells[1] if len(cells) > 1 else ""
        if not field_name or field_name in ("字段", ""):
            continue

        field_info = {
            "field": field_name,
            "meaning": cells[2] if len(cells) > 2 else "",
            "cycle": cells[3] if len(cells) > 3 else "",
            "dict_value": cells[4] if len(cells) > 4 else "",
            "note": cells[5] if len(cells) > 5 else ""
        }
        current_fields.append(field_info)

    # 保存最后一个分类
    if current_fields or current_category:
        result["sections"].append({
            "category": current_category or "基础字段",
            "fields": current_fields
        })

    # 解析口径案例（从案例指标标题之后）
    case_started = False
    for i in range(header_line_idx + 1, len(lines)):
        line = lines[i].strip()

        if re.search(r"案例指标", line):
            case_started = True
            continue
        if not case_started:
            continue
        if not line.startswith("|"):
            continue
        if "---" in line:
            continue

        cells = [c.strip() for c in line.split("|")[1:-1]]
        if len(cells) < 2:
            continue

        name = cells[0] if cells[0] else ""
        sql_cell_idx = -1
        for j, c in enumerate(cells[1:], 1):
            if c and any(k in c for k in ["=", "and", "or", "date_format", "COALESCE", "prod_type", "is_"]):
                sql_cell_idx = j
                break

        if sql_cell_idx == -1 and not name:
            continue

        if not name and sql_cell_idx >= 0:
            if result["case_metrics"] and cells[sql_cell_idx] if sql_cell_idx < len(cells) else "":
                last_metric = result["case_metrics"][-1]
                sql_text = cells[sql_cell_idx] if sql_cell_idx < len(cells) else ""
                last_metric["sql"] = last_metric["sql"].rstrip() + " " + sql_text
                continue

        sql = cells[sql_cell_idx] if sql_cell_idx >= 0 and sql_cell_idx < len(cells) else ""

        if name or sql:
            result["case_metrics"].append({
                "name": name,
                "sql": sql
            })


def _parse_sections_without_category(lines: list, result: dict):
    """解析无字段分类的格式（002/003/005 类型）：只有字段名/字段含义/注释 3列
    可能的表头变体：
    - 002: 字段名 / 字段含义 / 注释
    - 005/003: col_name / data_type / 注释（英文表头）
    """
    # 找到表头行（支持多种变体）
    header_line_idx = -1
    for i, line in enumerate(lines):
        if re.search(r"字段名|字段类|col_name|^字段$", line):
            header_line_idx = i
            break

    # 如果没有找到标准表头，尝试检测无表头行的两列格式
    if header_line_idx == -1:
        for i, line in enumerate(lines):
            if not line.strip().startswith("|"):
                continue
            cells = [c.strip() for c in line.split("|")[1:-1]]
            # 两列格式：字段名在第0列，含义在第1列，且都不是数据类型关键字
            if len(cells) == 2 and cells[0] and cells[1]:
                if cells[0] not in ("字段", "字段名", "col_name", "字段分类", "指标名称", "string", "bigint", "varchar", "decimal", "timestamp", "date", "int", "float", "double"):
                    header_line_idx = i
                    break
        if header_line_idx == -1:
            logger.warning("未找到字段表头（无分类格式），跳过字段解析")
            return

    # 解析字段（所有字段归入"基础字段"分类）
    current_fields = []

    for i in range(header_line_idx + 1, len(lines)):
        line = lines[i].strip()

        if not line or line.startswith("| ---"):
            continue
        if not line.startswith("|"):
            continue

        cells = [c.strip() for c in line.split("|")[1:-1]]
        if len(cells) < 2:
            continue

        # 检测空第一列格式：某些 Excel 表格第一列为空，数据从 B 列开始
        # 004格式: cells[0]="", cells[1]=field_name, cells[2]=data_type, cells[3]=meaning
        # 014格式: cells[0]="", cells[1]=field_name, cells[2]="", cells[3]=source_table
        # 通过 cells[2] 是否为数据类型来区分
        data_types = ("string", "bigint", "varchar", "decimal", "timestamp", "date", "int", "float", "double")
        is_004_format = (cells[0] == "" and len(cells) > 3 and
                          cells[2] in data_types)
        is_014_format = (cells[0] == "" and len(cells) > 3 and
                          cells[2] == "" and cells[3] != "")

        if is_004_format:
            # 004 格式：col_name / data_type / 注释 / 中文含义
            field_name = cells[1]
            if not field_name or field_name in ("字段", "col_name", ""):
                continue
            field_info = {
                "field": cells[1],
                "meaning": cells[3] if len(cells) > 3 else "",
                "cycle": "",
                "dict_value": "",
                "note": cells[2] if len(cells) > 2 else ""
            }
            current_fields.append(field_info)
            continue
        elif is_014_format:
            # 014 格式：col_name / "" / col_meaning / source_table（3+列）
            field_name = cells[1]
            if not field_name or field_name in ("字段", "col_name", ""):
                continue
            field_info = {
                "field": cells[1],
                "meaning": cells[2] if len(cells) > 2 else "",
                "cycle": "",
                "dict_value": "",
                "note": cells[3] if len(cells) > 3 else ""
            }
            current_fields.append(field_info)
            continue

        # 对于空首列格式，字段名在 cells[1]
        # 对于空首列格式但不符合 004/014 的情况，使用 014 格式逻辑
        # 014 格式: cells[0]="", cells[1]=field_name, cells[2]="", cells[3]=source
        if cells[0] == "" and len(cells) > 1 and cells[1]:
            field_name = cells[1]
            if field_name not in ("字段", "col_name", ""):
                field_info = {
                    "field": cells[1],
                    "meaning": cells[2] if len(cells) > 2 else "",
                    "cycle": "",
                    "dict_value": "",
                    "note": cells[3] if len(cells) > 3 else ""
                }
                current_fields.append(field_info)
            continue

        field_name = cells[0]
        if not field_name or field_name in ("字段", "col_name", ""):
            continue

        # 判断是 005 类型（英文表头）还是 002 类型（中文表头）
        # 005: 字段在第1列，类型在第2列，注释在第3列
        # 002: 字段在第1列，含义在第2列，注释在第3列
        # 通过 cells[1] 是否为数据类型（string/bigint/varchar等）来判断
        is_english_header = cells[1] in ("string", "bigint", "varchar", "decimal", "timestamp", "date", "int", "float", "double")

        if is_english_header:
            # 005 格式：col_name / data_type / 中文注释
            field_info = {
                "field": field_name,
                "meaning": cells[2] if len(cells) > 2 else "",
                "cycle": "",
                "dict_value": "",
                "note": ""
            }
        else:
            # 002 格式：字段名 / 字段含义 / 注释（2-3列）
            field_info = {
                "field": field_name,
                "meaning": cells[1] if len(cells) > 1 else "",
                "cycle": "",
                "dict_value": "",
                "note": cells[2] if len(cells) > 2 else ""
            }
        current_fields.append(field_info)

    if current_fields:
        result["sections"].append({
            "category": "基础字段",
            "fields": current_fields
        })


def build_output_md(data: dict) -> str:
    """将解析后的数据构建为最终 md 格式"""
    lines = []

    # Header
    lines.append(f"# {data['table_name']}\n")

    if data["hive_table"]:
        lines.append(f"- **Hive 表名**：`{data['hive_table']}`")
    if data["view_name"]:
        lines.append(f"- **视图名**：`{data['view_name']}`")

    lines.append("\n---\n")

    # 字段说明
    lines.append("## 字段说明\n")

    for section in data["sections"]:
        category = section["category"] if section["category"] else "基础字段"
        lines.append(f"### {category}\n")

        if not section["fields"]:
            lines.append("*（无字段）*\n")
            continue

        lines.append("| 字段 | 字段含义 | 标签周期 | 字典值 | 说明 |")
        lines.append("|------|---------|---------|-------|------|")

        for f in section["fields"]:
            field = f.get("field", "")
            meaning = f.get("meaning", "")
            cycle = f.get("cycle", "")
            dict_value = f.get("dict_value", "")
            note = f.get("note", "")

            # 转义 md 特殊字符
            def esc(s):
                return str(s).replace("|", "\\|").replace("\n", "<br>")

            lines.append(f"| {esc(field)} | {esc(meaning)} | {esc(cycle)} | {esc(dict_value)} | {esc(note)} |")

        lines.append("")

    # 口径案例
    if data["case_metrics"]:
        lines.append("---\n")
        lines.append("## 口径案例\n")
        lines.append("| 指标名称 | 计算口径 |")
        lines.append("|---------|---------|")

        for metric in data["case_metrics"]:
            name = metric.get("name", "")
            sql = metric.get("sql", "").replace("|", "\\|")
            lines.append(f"| {name} | `{sql}` |")

        lines.append("")

    return "\n".join(lines)


def get_table_name_from_path(excel_path: str) -> str:
    """从 Excel 文件路径提取表名作为输出文件名"""
    basename = os.path.splitext(os.path.basename(excel_path))[0]
    return basename


def convert_excel_to_table_md(excel_path: str, output_dir: str = None) -> str:
    """主转换函数"""
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

    logger.info(f"开始转换: {excel_path}")

    # Step 1: 读取 Excel 并转为中间 md
    logger.info("Step 1: 读取 Excel...")
    intermediate_md = read_excel_to_md(excel_path)

    # Step 2: 解析中间 md
    logger.info("Step 2: 解析中间 md...")
    data = parse_intermediate_md(intermediate_md)
    logger.info(f"  - 表名: {data['table_name']}")
    logger.info(f"  - Hive表名: {data['hive_table']}")
    logger.info(f"  - 视图名: {data['view_name']}")
    logger.info(f"  - 字段分类数: {len(data['sections'])}")
    if data["sections"]:
        for s in data["sections"]:
            logger.info(f"    - {s['category']}: {len(s['fields'])} 字段")
    logger.info(f"  - 口径案例数: {len(data['case_metrics'])}")

    # Step 3: 构建输出 md
    logger.info("Step 3: 构建最终 md...")
    output_md = build_output_md(data)

    # Step 4: 写入输出文件
    output_dir = output_dir or REFERENCES_DIR
    os.makedirs(output_dir, exist_ok=True)

    table_name_for_file = get_table_name_from_path(excel_path)
    output_path = os.path.join(output_dir, f"{table_name_for_file}.md")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write(output_md)

    logger.info(f"完成，输出文件: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(description="Excel to Table MD")
    parser.add_argument("excel_path", help="Excel 文件路径")
    parser.add_argument("-o", "--output-dir", help="输出目录（默认：references 目录）")
    parser.add_argument("-v", "--verbose", action="store_true", help="详细输出")
    args = parser.parse_args()

    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    output_path = convert_excel_to_table_md(
        args.excel_path,
        output_dir=args.output_dir
    )
    print(f"\n输出文件: {output_path}")


if __name__ == "__main__":
    main()
